import hashlib
import time
import json
import os

import numpy as np
from PIL import Image
from io import BytesIO
from wx_sdk import wx_sdk
import openpyxl


def get_json_value_by_key(in_json, target_key, results=[]):
    """字典,列表,元组的多类型嵌套"""
    if isinstance(in_json, dict):  # 如果输入数据的格式为dict
        for key in in_json.keys():  # 循环获取key
            data = in_json[key]
            get_json_value_by_key(data, target_key, results=results)  # 回归当前key对于的value

            if key == target_key:  # 如果当前key与目标key相同就将当前key的value添加到输出列表
                results.append(data)

    elif isinstance(in_json, list) or isinstance(in_json, tuple):  # 如果输入数据格式为list或者tuple
        for data in in_json:  # 循环当前列表
            get_json_value_by_key(data, target_key, results=results)  # 回归列表的当前的元素
    return results


def change_img(file, max_width=2048, max_height=2048, min_width=256, min_height=256):
    image = Image.open(file)  # 读取图片
    w, h = image.size
    if min_width <= w <= max_width and min_height <= h <= max_height:
        # print('is OK.')
        scale = 1
        new_img = image
    elif (1.0 * w / max_width) > (1.0 * h / max_height):
        scale = 1.0 * w / max_width
        new_img = image.resize((int(w / scale), int(h / scale)), Image.ANTIALIAS)
        # print((int(w / scale), int(h / scale)))
    else:
        scale = 1.0 * h / max_height
        new_img = image.resize((int(w / scale), int(h / scale)), Image.ANTIALIAS)
        # print((int(w / scale), int(h / scale)))

    output = BytesIO()
    new_img.save(output, format='png')
    contents = output.getvalue()
    output.close()
    return contents, scale


def cosine_similarity(x, y, norm=False):
    """ 计算两个向量x和y的余弦相似度 """
    assert len(x) == len(y), "len(x) != len(y)"
    zero_list = [0] * len(x)
    if x == zero_list or y == zero_list:
        return float(1) if x == y else float(0)

    # method 1
    res = np.array([[x[i] * y[i], x[i] * x[i], y[i] * y[i]] for i in range(len(x))])
    cos = sum(res[:, 0]) / (np.sqrt(sum(res[:, 1])) * np.sqrt(sum(res[:, 2])))
    return 0.5 * cos + 0.5 if norm else cos  # 归一化到[0, 1]区间内


def algorhtmReq(img_file):
    url = 'https://aiapi.jd.com/jdai/personreid'
    timestamp = int(time.time() * 1000)
    secretkey = '8569938395247b228a1e356bf40d89bc'

    sign_str = secretkey + str(timestamp)
    hl = hashlib.md5()
    hl.update(sign_str.encode(encoding='utf-8'))
    sign = hl.hexdigest()

    params = {
        'muti_det': '2',
        'appkey': '30598bbc33f844e4cf33bfc77181a616',
        'timestamp': timestamp,
        'sign': sign
    }

    # image size must be between 256X256 and 4096X4096
    # image size exceeds 2M
    contents, scale = change_img(img_file)

    # response = wx_sdk.wx_post_req(url, params, img=img_file)
    response = wx_sdk.wx_post_req(url, params, bodyStr=contents)
    datas = json.loads(response.text)
    # print(datas)

    # 构建统一格式
    # attributes
    attributes = get_json_value_by_key(datas, 'personreidResult', results=[])[0]
    # print(attributes)

    # 位置信息--x1,y1,x2,y2,score分别表示预测出的人体矩形框的左上角坐标，右下角坐标，以及置信度。
    # 返归的矩形框按照置信度高低排序，当未检测到人体时，返回空array[]
    dict_value = get_json_value_by_key(datas, 'humanDetectionResult', results=[])[0]
    # print(dict_value)  # [[317, 16, 785, 1399, 0.9080960154533386]]
    result_list = []
    # 识别人数
    total_dict = {'total_num': len(dict_value)}
    result_list.append(total_dict)

    for i in range(len(dict_value)):
        result_dict = {
            "position": [int(j/scale) for j in dict_value[i][:4]],
            "score": dict_value[i][-1],
            'attributes': attributes[i]
        }
        result_list.append(result_dict)

    # print(result_list)
    return result_list


def compared_image_url(original_img, resources_img):
    """余弦相似度对比"""
    # 获得图片特征矩阵
    # 假如初始图片为1个人
    original_result = algorhtmReq(original_img)
    # 对比图
    resources_result = algorhtmReq(resources_img)
    score_list = []
    for i in range(resources_result[0]["total_num"]):
        # 余弦相似度
        score = cosine_similarity(original_result[1]['attributes'], resources_result[i+1]['attributes'])
        score_list.append(score)
    return score_list


def compared_image_json(original_img, resources_img):
    """余弦相似度对比"""
    # 获得图片特征矩阵
    with open("attributes_dict.json", "r") as f:
        attributes_dict = f.read()
        attributes_dict = json.loads(attributes_dict)
    # 假如初始图片为1个人
    original_result = attributes_dict[original_img.split(".")[0]]
    # 对比图
    resources_result = attributes_dict[resources_img.split(".")[0]]
    score_list = []
    for i in range(resources_result[0]["total_num"]):
        # 余弦相似度
        score = cosine_similarity(original_result[1]['attributes'], resources_result[i + 1]['attributes'])
        score_list.append(score)
    return score_list


def save_attributes(image_path):
    # image_path = "/Users/tao/Desktop/testAI/similar_images/"
    image_list = os.listdir(image_path)
    if ".DS_Store" in image_list:
        image_list.remove(".DS_Store")
    if "Thumbs.db" in image_list:
        image_list.remove("Thumbs.db")
    image_list.sort()
    attributes_dict = {}
    for img in image_list[:100]:
        print(img)
        image_index = img.split(".")[0]
        attributes = algorhtmReq(os.path.join(image_path, img))
        attributes_dict[image_index] = attributes
        time.sleep(1)
    # print(attributes_dict)
    attributes_dict = json.dumps(attributes_dict)
    save_path = "query_attributes_dict_100.json"
    with open(save_path, "w") as f:
        f.write(attributes_dict)
    print("记录完成")


if __name__ == '__main__':
    # path = "/Users/tao/Desktop/testAI/similar_images/"
    path = "/Users/tao/Desktop/Person-Attribute-Recognition-MarketDuke/dataset/Market-1501/query"

    # # 保存attributes
    save_attributes(path)

    # # excel
    # img_list = os.listdir(path)
    # if ".DS_Store" in img_list:
    #     img_list.remove(".DS_Store")
    # img_list.sort()
    # # wb = openpyxl.load_workbook("十人对比.xlsx")
    # # sh = wb["Sheet2"]
    #
    # for i in img_list[:]:
    #     row = img_list.index(i) + 2
    #     # sh.cell(row=row, column=1, value=i.split(".")[0])
    #     for j in img_list:
    #         column = img_list.index(j) + 2
    #         # sh.cell(row=1, column=column, value=j.split(".")[0])
    #         # print(i)
    #         # print(j)
    #         result = compared_image_json(i, j)
    #         # sh.cell(row=row, column=column, value=str(result[0]))
    #         print(row, column, result)
    # # wb.close()
    # # wb.save("十人对比.xlsx")

    # img1 = "/Users/tao/Desktop/testAI/similar_images/0194_c1s1_060031_01.jpg"  # 样本图片
    # img2 = "/Users/tao/Desktop/testAI/similar_images/0194_c1s1_060056_01.jpg"  # 和样本对比的图片
    # result = compared_image_url(img1, img2)
    # print(result)  # 0.9672974181515218  0.9569536908667715  0.9614792827124997
